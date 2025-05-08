from fastapi import FastAPI, APIRouter, Depends, HTTPException,UploadFile, File
import requests
from schema.appointment import Appointment, AppointmentCreate
from db.config import get_db
from fastapi import status
from dotenv import load_dotenv
import os
import json
from datetime import datetime
import openpyxl
from motor.motor_asyncio import AsyncIOMotorDatabase
from io import BytesIO
from typing import List
load_dotenv()



dentally_router = APIRouter()


Dentally_appointments_url = "https://api.dentally.co/v1/appointments"
Dentally_practitioners_url = "https://api.dentally.co/v1/practitioners"
Dentally_patient_url = "https://api.dentally.co/v1/patients"
Dentally_payment_plans_url = "https://api.dentally.co/v1/payment_plans?active=true"
Dentally_api_key = os.getenv("DENTALLY_API_KEY")

def validate_date_format(date_str: str):
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Date must be in YYYY-MM-DD format")
    

@dentally_router.get("/fetch-all-dentally-appointments/{date}", status_code=status.HTTP_200_OK)
async def get_all_dentally_appointments(
    date: str,
    db: AsyncIOMotorDatabase = Depends(get_db)
):
    validated_date = validate_date_format(date)
    headers = {"Authorization": f"Bearer {Dentally_api_key}"}

    all_appointments = []
    page = 1
    per_page = 100

    while True:
        url = f"{Dentally_appointments_url}?on={validated_date}&page={page}&per_page={per_page}"
        response = requests.get(url, headers=headers)

        if response.status_code != 200:
            raise HTTPException(
                status_code=response.status_code,
                detail=f"Error fetching appointments on page {page}: {response.text}"
            )

        data = response.json()
        appointments = data.get("appointments", [])
        meta = data.get("meta", {})

        all_appointments.extend(appointments)

        if meta.get("current_page") >= meta.get("total_pages", 1):
            break
        page += 1

    # Save appointments to MongoDB
    await db["appointments"].delete_many({})
    if all_appointments:
        await db["appointments"].insert_many(all_appointments)
        return {
            "message": f"Fetched and saved {len(all_appointments)} appointments from {meta.get('total_pages')} page(s)."
        }
    else:
        return {"message": "No appointments found to save"}


def set_nested_value(dictionary, keys, value):
    for key in keys[:-1]:
        dictionary = dictionary.setdefault(key, {})
    dictionary[keys[-1]] = value

@dentally_router.post("/upload-practitioners-excel-file/")
async def upload_practiner_excel(file: UploadFile = File(...), db: AsyncIOMotorDatabase = Depends(get_db)):
    try:
        contents = await file.read()
        workbook = openpyxl.load_workbook(BytesIO(contents))
        sheet = workbook.active
        await db["practitioners"].delete_many({})
        headers = [cell.value for cell in sheet[1]]
        data_to_insert = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            nested_record = {}
            for key, value in zip(headers, row):
                if not key:
                    continue
                keys = key.split(".")
                set_nested_value(nested_record, keys, value)
            data_to_insert.append(nested_record)

        if not data_to_insert:
            raise HTTPException(status_code=400, detail="No data to insert")

        db["practitioners"].insert_many(data_to_insert)
        return {"message": "Upload successful", "inserted_count": len(data_to_insert)}

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@dentally_router.post("/upload-practitioners-mapping-excel-file/")
async def upload_mapping_excel_file(
    file: UploadFile = File(...),
    db: AsyncIOMotorDatabase = Depends(get_db)
):
    try:
        contents = await file.read()
        workbook = openpyxl.load_workbook(BytesIO(contents))
        sheet = workbook.active

        # Extract headers from the first row
        headers = [cell.value for cell in sheet[1]]
        if not headers or None in headers:
            raise HTTPException(status_code=400, detail="Invalid or missing headers in the Excel file.")

        await db["treatment_details"].delete_many({})  # Optional: Clear existing data

        data_to_insert = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue  # Skip completely empty rows

            record = dict(zip(headers, row))
            # Optional cleaning: trim whitespace and ensure string conversion where needed
            cleaned_record = {
                str(k).strip(): (str(v).strip() if isinstance(v, str) else v)
                for k, v in record.items()
            }
            data_to_insert.append(cleaned_record)

        if not data_to_insert:
            raise HTTPException(status_code=400, detail="No valid data to insert.")

        await db["treatment_details"].insert_many(data_to_insert)

        return {
            "message": "Upload successful",
            "inserted_count": len(data_to_insert)
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")



@dentally_router.post("/sync/payment-plans")
async def sync_payment_plans(db: AsyncIOMotorDatabase = Depends(get_db)):
    try:
        # Fetch data from Dentally API
        headers = {"Authorization": f"Bearer {Dentally_api_key}"}
        response = requests.get(Dentally_payment_plans_url, headers=headers)
        response.raise_for_status()
        data = response.json()

        payment_plans = data.get("payment_plans", [])
        if not payment_plans:
            raise HTTPException(status_code=404, detail="No payment plans found in API response")

        # Optional: Clear old data
        await db["payment_plans"].delete_many({})

        # Insert each plan into MongoDB
        await db["payment_plans"].insert_many(payment_plans)

        return {
            "message": "✅ Payment plans fetched and stored successfully.",
            "stored_count": len(payment_plans)
        }

    except requests.RequestException as e:
        raise HTTPException(status_code=502, detail=f"❌ Failed to fetch payment plans: {e}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"❌ Unexpected error: {e}")
    

def get_availability_from_dentally(practitioner_ids: List[int], start_time: str, finish_time: str,duration:int=60):

    url = "https://api.dentally.co/v1/appointments/availability"
    headers = {"Authorization": f"Bearer {Dentally_api_key}"}
    
    params = {
        "practitioner_ids[]": [practitioner_ids],  # Pass as array
        "start_time": start_time,
        "duration": duration,
        "finish_time": finish_time,
    }
    
    response = requests.get(url, params=params,headers=headers)
    print(response.content)
    if response.status_code == 200:
        return response.json()
    else:
        print(f"Error fetching availability: {response.status_code} - {response.text}")
        return []